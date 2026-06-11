"""
WFM Falabella — Backend FastAPI
"""
from fastapi import FastAPI, UploadFile, Form, File
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd, numpy as np, io, json, uuid
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from modelo import proyectar_mes, COLA_INFO, VENTURES_SIN_META

app = FastAPI(title="WFM Falabella API")

app.add_middleware(CORSMiddleware,
    allow_origins=["http://localhost:3000","http://localhost:5173"],
    allow_methods=["*"], allow_headers=["*"])

_resultados = {}

@app.get("/health")
def health(): return {"status":"ok"}

@app.post("/calcular")
async def calcular(
    file_hist: UploadFile = File(...),
    anio:  str = Form(...),
    mes:   str = Form(...),
    metas: str = Form(...),
    erlang:str = Form(...)
):
    try:
        anio_i  = int(anio)
        mes_i   = int(mes)
        metas_d = json.loads(metas)
        ep_d    = json.loads(erlang)

        content = await file_hist.read()
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb['Historico']
        rows = list(ws.iter_rows(values_only=True))
        df = pd.DataFrame(rows[1:], columns=['Fecha','Dia','Hora','Cola','Entrantes','TMO'])
        df = df.dropna(subset=['Fecha','Cola'])
        df['Fecha']    = pd.to_datetime(df['Fecha'])
        df['Entrantes']= pd.to_numeric(df['Entrantes'], errors='coerce').fillna(0)
        df['año']      = df['Fecha'].dt.year
        df['mes']      = df['Fecha'].dt.month
        df['dow']      = df['Fecha'].dt.dayofweek
        df['Hora']     = pd.to_numeric(df['Hora'], errors='coerce').fillna(0).astype(int)
        df = df[df['Cola'].isin(COLA_INFO.keys())]

        df_proj = proyectar_mes(df, anio_i, mes_i, metas_d, ep_d)
        if df_proj.empty:
            return {"error": "No se generaron datos. Verifica el histórico."}

        sid = str(uuid.uuid4())
        _resultados[sid] = {'df': df_proj, 'anio': anio_i, 'mes': mes_i}

        resumen = {}
        for v, grp in df_proj.groupby('Venture'):
            resumen[v] = {
                'total':    round(float(grp['Proyeccion'].sum())),
                'ftes_max': int(grp['Ftes'].max()),
            }

        return {"session_id": sid, "resumen": resumen, "total_filas": len(df_proj)}

    except Exception as e:
        import traceback
        return {"error": str(e) + "\n" + traceback.format_exc()}


@app.get("/descargar/{sid}")
def descargar(sid: str):
    if sid not in _resultados:
        return {"error": "Sesión expirada"}

    data    = _resultados[sid]
    df_proj = data['df']
    anio    = data['anio']
    mes     = data['mes']
    mes_nombre = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',
                  6:'Junio',7:'Julio',8:'Agosto',9:'Septiembre',
                  10:'Octubre',11:'Noviembre',12:'Diciembre'}[mes]

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Proyeccion"
    ws_out.freeze_panes = 'A2'

    MED = "2F5496"; WHITE = "FFFFFF"; GRAY = "F5F7FA"
    thin_s = Side(style='thin', color='E2E8F0')
    thin = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    headers = ['Date','Hour','Canal','Proyección','Ftes','Proyección ajustada',
               'Country','Grupo','Venture','Contact center','Contact center Secundario']
    col_widths = [14,8,12,18,8,20,10,12,18,22,52]

    for ci,(h,w) in enumerate(zip(headers,col_widths),1):
        c = ws_out.cell(1,ci,h)
        c.font = Font(name='Arial',bold=True,color=WHITE,size=10)
        c.fill = PatternFill('solid',start_color=MED)
        c.alignment = Alignment(horizontal='center',vertical='center')
        c.border = thin
        ws_out.column_dimensions[get_column_letter(ci)].width = w
    ws_out.row_dimensions[1].height = 20

    def country(cola):
        for s,c in [('_CH','CH'),('_PE','PE'),('_CO','CO'),('_AR','AR'),('_UY','UY')]:
            if cola.endswith(s): return c
        return ''

    def grupo(v):
        if v in ('FACL','FACO','FAPE'): return 'F.com'
        if any(x in v for x in ('SOC','SOPE','TIENDAS','TS')): return 'Sodimac'
        if any(x in v for x in ('TOCL','TOPE')): return 'Tottus'
        return 'Otro'

    df_s = df_proj.sort_values(['Cola','Fecha','Hora'])
    for ri,(_, r) in enumerate(df_s.iterrows()):
        row = ri+2
        bg = WHITE if ri%2==0 else GRAY
        vals = [r['Fecha'].strftime('%Y-%m-%d'), int(r['Hora']), r['Canal'],
                round(float(r['Proyeccion']),2), int(r['Ftes']), '',
                country(r['Cola']), grupo(r['Venture']), r['Venture'], '', r['Cola']]
        for ci,v in enumerate(vals,1):
            c = ws_out.cell(row,ci,v)
            c.font = Font(name='Arial',size=9)
            c.fill = PatternFill('solid',start_color=bg)
            c.alignment = Alignment(horizontal='center' if ci!=11 else 'left',vertical='center')
            c.border = thin
        ws_out.row_dimensions[row].height = 13

    # Hoja resumen
    ws_res = wb_out.create_sheet("Resumen")
    ws_res.sheet_view.showGridLines = False
    for ci,(h,w) in enumerate(zip(['Venture','Canal','Total Entrantes','FTEs Prom','FTEs Máx'],
                                   [20,12,18,14,12]),1):
        c = ws_res.cell(1,ci,h)
        c.font = Font(name='Arial',bold=True,color=WHITE,size=10)
        c.fill = PatternFill('solid',start_color="1F3864")
        c.alignment = Alignment(horizontal='center',vertical='center')
        c.border = thin
        ws_res.column_dimensions[get_column_letter(ci)].width = w

    df_r = df_proj.groupby(['Venture','Canal']).agg(
        total=('Proyeccion','sum'), avg=('Ftes','mean'), mx=('Ftes','max')
    ).reset_index()
    for ri2,r2 in enumerate(df_r.itertuples(),2):
        bg2 = WHITE if ri2%2==0 else GRAY
        for ci,v in enumerate([r2.Venture,r2.Canal,round(r2.total),round(r2.avg,1),int(r2.mx)],1):
            c = ws_res.cell(ri2,ci,v)
            c.font = Font(name='Arial',size=10)
            c.fill = PatternFill('solid',start_color=bg2)
            c.alignment = Alignment(horizontal='center',vertical='center')
            c.border = thin

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)

    return StreamingResponse(buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition':f'attachment; filename=Proyeccion_{mes_nombre}_{anio}.xlsx'})
